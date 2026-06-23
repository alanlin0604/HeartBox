"""One-shot: convert the v3 synthetic training CSVs into xlsx with the
same 3-sheet format the export_ml_training_data management command
produces, so the defense slide can show the FULL 22,796-row training
corpus rather than the 43-row demo-DB subset.

Run from backend/ (so venv pandas/openpyxl are on path):
    .\venv\Scripts\python.exe ml\scripts\csv_to_xlsx_for_defense.py
"""
import sys
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment


HERE = Path(__file__).resolve().parents[1]
DATASETS = HERE / 'datasets'

LAG_WINDOWS = (1, 3, 7, 14)

# Mirror of the description dict from export_ml_training_data.py — kept
# in sync manually since this one-shot script is for the defense slide
# only, not part of the regular pipeline.
DESC_TEMPLATE = {
    'sent_lag_{w}d_mean':          ('近 {w} 天平均情緒分數（-1 至 +1）',         'MoodNote.sentiment_score'),
    'stress_lag_{w}d_mean':        ('近 {w} 天平均壓力指數（0-10）',             'MoodNote.stress_index'),
    'stress_lag_{w}d_max':         ('近 {w} 天最高壓力指數',                     'MoodNote.stress_index'),
    'entries_lag_{w}d':            ('近 {w} 天日記則數',                         'MoodNote 計數'),
    'sleep_lag_{w}d_mean':         ('近 {w} 天平均睡眠時數',                     'DailySleep.sleep_hours'),
    'sleep_quality_lag_{w}d_mean': ('近 {w} 天平均主觀睡眠品質（1-5）',           'DailySleep.sleep_quality'),
    'deep_sleep_pct_lag_{w}d_mean':('近 {w} 天深層睡眠占比',                     'DailySleep.deep/light/rem 分鐘'),
    'habit_lag_{w}d_mean':         ('近 {w} 天習慣完成率（0-1）',                'HabitLog ÷ 活躍 Habit 數'),
    'steps_lag_{w}d_mean':         ('近 {w} 天平均步數',                         'HealthMetric (steps)'),
    'exercise_lag_{w}d_mean':      ('近 {w} 天平均運動分鐘',                     'HealthMetric (exercise_minutes)'),
    'hrv_lag_{w}d_mean':           ('近 {w} 天平均心率變異 (HRV)',                'HealthMetric (hrv)'),
    'rhr_lag_{w}d_mean':           ('近 {w} 天平均靜止心率',                     'HealthMetric (heart_rate)'),
}
STATIC_DESC = {
    'user_id':         ('使用者 ID（已匿名化）',                                'CustomUser.id'),
    'ref_date':        ('該樣本的參考日（features 為當天回溯）',                 '推導自 MoodNote.created_at'),
    'day_of_week':     ('星期幾（0=週一 … 6=週日）',                            '推導自 ref_date'),
    'is_weekend':      ('是否週末（0/1）',                                       '推導自 ref_date'),
    'day_of_month':    ('幾號（1-31）',                                          '推導自 ref_date'),
    'current_streak':  ('目前連續寫日記天數',                                    'JournalStreak.current_streak'),
    'bedtime_std_14d': ('近 14 天就寢時間標準差（生活規律性指標）',              'DailySleep.bedtime'),
    'target_sentiment':('預測標的：第 horizon 天的情緒分數',                     '同 sent_lag 但取未來日'),
    'target_stress':   ('預測標的：第 horizon 天的壓力指數',                     '同 stress_lag 但取未來日'),
    'target_spike':    ('預測標的：未來 horizon 天內出現高壓事件（stress>=7）為 1', 'MoodNote.stress_index'),
}


def describe_column(col):
    if col in STATIC_DESC:
        return STATIC_DESC[col]
    for tpl_key, (tpl_cn, tpl_src) in DESC_TEMPLATE.items():
        for w in LAG_WINDOWS:
            if col == tpl_key.format(w=w):
                return tpl_cn.format(w=w), tpl_src
    return '—', '—'


def build_stats_rows(task, df, n_users, n_rows, feature_cols, target_cols):
    rows = [
        {'項目': '任務 (task)',                       '值': task},
        {'項目': '資料來源',                          '值': '合成資料（v3 synthetic, 模擬真實使用者行為）'},
        {'項目': '總樣本數 (rows)',                   '值': n_rows},
        {'項目': '不同使用者數 (users)',              '值': n_users},
        {'項目': 'feature 欄位數',                    '值': len(feature_cols)},
        {'項目': 'target 欄位數',                     '值': len(target_cols)},
        {'項目': '時間窗口 (lag windows)',            '值': '1 / 3 / 7 / 14 天'},
        {'項目': '預測 horizon',                      '值': '3 天'},
    ]
    if task == 'mood_prediction':
        s = df['target_sentiment'].dropna()
        rows.append({'項目': 'target_sentiment min/mean/max',
                     '值': f'{s.min():.3f} / {s.mean():.3f} / {s.max():.3f}'})
        if 'target_stress' in df.columns:
            s = df['target_stress'].dropna()
            rows.append({'項目': 'target_stress min/mean/max',
                         '值': f'{s.min():.2f} / {s.mean():.2f} / {s.max():.2f}'})
    else:
        pos = int(df['target_spike'].sum())
        rows.append({'項目': '正樣本數 (高壓事件 = 1)', '值': f'{pos} ({pos / n_rows * 100:.1f}%)'})
        rows.append({'項目': '負樣本數 (無高壓 = 0)',    '值': f'{n_rows - pos} ({(n_rows - pos) / n_rows * 100:.1f}%)'})
    return rows


def convert(task, csv_path, xlsx_path):
    print(f'Loading {csv_path.name} ...')
    df = pd.read_csv(csv_path)
    n_rows = len(df)
    n_users = df['user_id'].nunique() if 'user_id' in df.columns else 0
    feature_cols = [c for c in df.columns if c not in ('user_id', 'ref_date') and not c.startswith('target_')]
    target_cols = [c for c in df.columns if c.startswith('target_')]
    print(f'  rows={n_rows}, users={n_users}, features={len(feature_cols)}, targets={len(target_cols)}')

    desc_rows = [{'欄位': c, '中文說明': cn, '資料來源': src}
                 for c in df.columns
                 for cn, src in [describe_column(c)]]
    df_desc = pd.DataFrame(desc_rows)
    df_stats = pd.DataFrame(build_stats_rows(task, df, n_users, n_rows, feature_cols, target_cols))

    with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Raw_Features', index=False)
        df_desc.to_excel(writer, sheet_name='Feature_Description', index=False)
        df_stats.to_excel(writer, sheet_name='Sample_Distribution', index=False)

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='C2410C')
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_cells in ws.columns:
                longest = max((len(str(c.value)) for c in col_cells if c.value is not None), default=0)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max(longest + 2, 10), 40)

    size_mb = xlsx_path.stat().st_size / (1024 * 1024)
    print(f'  → wrote {xlsx_path.name} ({size_mb:.1f} MB)')


def main():
    targets = [
        ('mood_prediction', DATASETS / 'mood_prediction_v3_synthetic.csv',
                            DATASETS / 'mood_prediction_FULL_22796rows.xlsx'),
        ('stress_spike',    DATASETS / 'stress_spike_v3_synthetic.csv',
                            DATASETS / 'stress_spike_FULL_31720rows.xlsx'),
    ]
    for task, csv_path, xlsx_path in targets:
        if not csv_path.exists():
            print(f'SKIP (missing): {csv_path}', file=sys.stderr)
            continue
        convert(task, csv_path, xlsx_path)


if __name__ == '__main__':
    main()
